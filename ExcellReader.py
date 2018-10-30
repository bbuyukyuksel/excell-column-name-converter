from openpyxl import load_workbook
from pprint import pprint
import json

class excellReader:
    row = None
    col = None
    _UniqueColumnElements = []
    _excellArray = []
    FROM = None
    TO = None

    def __init__(self,workbook):
        '''
        #Open Excell File
        ws = excellReader('./test.xlsx')

        #Get Data's ROWs x COLs
        rows, cols = ws.getExcellSize()
        print("Excell Sheet Size : [{}x{}]".format(rows, cols))

        #Create Index
        startRow = ws.createIndex('A', 1)
        endRow = ws.createIndex('C', rows)

        #Read Excell Array
        ws.readExcellArray(startRow, endRow)

        #Get Excell Array
        # excellArray = ws.getExcellArray()

        #Print Styled Array
        ws.styledArray()

        #Update Excell Array
        ws.setExcellArray(["Burak", "Büyükyüksel", "Can Sıkıntısı"], 0)
        ws.styledArray()

        #Save Changes to Excell File
        ws.saveExcellArray("myExcell.xlsx")
        '''
        self.wb = load_workbook(workbook)
        self.sheet = self.wb.active
        print(self.sheet)

    def readExcellArray(self,FROM,TO):
        self._excellArray = []
        self.FROM = FROM
        self.TO = TO
        _row = 0
        for cellObj in self.sheet[FROM:TO]:
            cols = []
            _col = 0
            for cell in cellObj:
                # print(cell.coordinate, cell.value)
                cols.append(cell.value)
                _col += 1
            self._excellArray.append(cols)
            _row += 1
            del cols
        self.row = _row
        self.col = _col
    def getExcellArray(self):
        return self._excellArray

    def getExcellArraySize(self):
        return (self.row,self.col)

    def setExcellArray(self,newLine,index):
        if type(newLine) == type(self._excellArray[index]) and len(newLine) == len(self._excellArray[index]):
            self._excellArray[index] = newLine
        else:
            print("##An Error Occured!")
            print("#Array        Type    :",type(newLine))
            print("#Excell Array Type    :",type(self._excellArray[index]))
            print("#Array        Length  :",len(newLine))
            print("#Excell Array Length  :",len(self._excellArray[index]))
            raise NameError("An Error Occured!")

    def saveExcellArray(self,name):
        row = 0
        #print("FROM:",self.FROM)
        #print("TO:",self.TO)
        for cellObj in self.sheet[self.FROM:self.TO]:
            #print("Row:",row)
            col = 0
            for cell in cellObj:
                #print("\tCol:",col)
                cell.value = self._excellArray[row][col]
                col += 1
            del col
            row += 1
        del row
        self.wb.save(name)

    def close(self):
        self.wb.close()

    def getExcellSize(self):
        return (self.sheet.max_row, self.sheet.max_column)

    def createIndex(self,ch,index):
        return '{}{}'.format(ch,str(index))

    def getUniqueColumnElements(self):
        self._UniqueColumnElements = []
        ROW, COL = self.row,self.col
        for i in range(ROW):
            for j in range(COL):
                if not self._excellArray[i][j] in self._UniqueColumnElements:
                    self._UniqueColumnElements.append(self._excellArray[i][j])
                else:
                    pass
        return self._UniqueColumnElements

    def getRenamedJSONFile(self, newName):
        name_format = newName + '{}'
        header = self._UniqueColumnElements[0]
        newNames = []
        for i in range(1, len(self._UniqueColumnElements)):
            # print("Old Name : ", UniqueColumnElements[i], end=" ->")
            # print("New Name : ", name_format.format(i+1))
            newNames.append(name_format.format(i))
        UpdatedNames = {header: {"oldNames": self._UniqueColumnElements[1:], "newNames": newNames}}
        return UpdatedNames

    def saveRenamedJSONFile(self,newName):
        UniqueColumnElements = self.getUniqueColumnElements()
        UpdatedNames = self.getRenamedJSONFile(newName)
        header = UniqueColumnElements[0]
        with open(header + '.json', 'w') as jsonfile:
            json.dump(UpdatedNames, jsonfile, indent=4)
        print("SAVED to {}".format(header))
        #pprint(UpdatedNames, indent=4)
    def styledArray(self):
        strExcellArr = str(self._excellArray)
        temp = strExcellArr[1:-1].replace('[','\n\t[')
        print("\033[01;31m", end="");
        print("Styled Excell Array :")
        print("\033[01;34m", end="");
        print(strExcellArr[0] + temp +'\n' + strExcellArr[-1])
        print("\033[00m", end="");